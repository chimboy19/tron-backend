
# # backend/utils/ocr_utils.py
# import tempfile
# import base64
# import os
# import re
# import logging
# from pdf2image import convert_from_bytes
# from config import Config
# from openai import OpenAI
# from difflib import get_close_matches
# import csv 


# client = OpenAI(api_key=Config.OPENAI_API_KEY) if Config.OPENAI_API_KEY else None

# log = logging.getLogger(__name__)

# def normalize_text(text: str) -> str:
#     if not text:
#         return ""
    
#     text = re.sub(r'[０-９]', lambda m: str(ord(m.group()) - 0xFEE0), text)
#     return re.sub(r'\s+', ' ', text.strip())

# def extract_items_from_text(text: str):
#     """Extract items (H-codes and quantities) from text."""
#     if not text:
#         return []
#     text = normalize_text(text)
#     items = []
    
#     hcod_pattern = re.compile(
#         r'H\s*(\d{6})\s*(?:[:：\-\s]*[x×*]?\s*(?:qty|数量|個|pcs|pc|ea|本|枚|セット|units?)?\s*[:：]?\s*)?(\d+)',
#         re.IGNORECASE
#     )
#     for m in hcod_pattern.finditer(text):
#         code = f"H{m.group(1)}"
#         qty = int(m.group(2))
#         items.append((code, qty))

#     all_hcods = re.findall(r'H\s*(\d{6})', text, re.IGNORECASE)
#     existing = {code for code, _ in items}
#     for digits in all_hcods:
#         code = f"H{digits}"
#         if code not in existing:
#             items.append((code, 1))

    
#     lines = re.split(r'[\n\r]+', text)
#     for line in lines:
#         line = line.strip()
#         if not line:
#             continue
#         part_qty = re.search(
#             r'([\w\s\-–―/／\(\)\[\]\u3000-\u9FFFΩμa-zA-Z0-9\.\+\=\&]{2,80}?)'
#             r'\s*[x×＊*]\s*'
#             r'(\d{1,5})'
#             r'(?:\s*(?:pcs|個|本|枚|セット|pc|ea|units?|点))?\b',
#             line,
#             re.IGNORECASE
#         )
#         if part_qty:
#             name = part_qty.group(1).strip()
#             qty = int(part_qty.group(2))
#             name = re.sub(r'[,\.\-_:\s;、。]+$', '', name)
#             if len(name) >= 2:
#                 items.append((name, qty))
#             continue

#         fallback = re.search(
#             r'([\w\s\-–―/／\(\)\[\]\u3000-\u9FFFΩμa-zA-Z0-9\.\+\=\&]{2,80}?)\s+(\d{1,4})\s*$',
#             line,
#             re.IGNORECASE
#         )
#         if fallback:
#             name = fallback.group(1).strip()
#             qty = int(fallback.group(2))
#             name = re.sub(r'[,\.\-_:\s;、。]+$', '', name)
#             if len(name) >= 2:
#                 items.append((name, qty))
#     return items



# def ocr_pdf_with_openai(pdf_bytes: bytes) -> str:
#     """OCR a PDF using OpenAI Vision API with Japanese text support."""
#     if not client:
#         log.warning("OpenAI client not configured; skipping OCR.")
#         return ""
    
#     try:
        
#         log.info("🔄 Converting PDF to image for OCR...")
#         images = convert_from_bytes(pdf_bytes, first_page=1, last_page=3)  
#         if not images:
#             log.warning("No images could be extracted from PDF")
#             return ""
        
#         all_ocr_text = ""
        
#         for page_num, image in enumerate(images, 1):
#             log.info(f"🔍 Processing page {page_num} for OCR...")
            
           
#             img_path = os.path.join(tempfile.gettempdir(), f"temp_ocr_page_{page_num}.jpg")
#             image.save(img_path, format="JPEG", quality=85)
            
#             try:
#                 with open(img_path, "rb") as f:
#                     encoded_img = base64.b64encode(f.read()).decode('utf-8')
            
#                 image_data_url = f"data:image/jpeg;base64,{encoded_img}"
                
#                 prompt = """
#                      make sure to avoid duplicates


#                     以下の注文書画像から、有効な「部品識別子」と「数量」のみを抽出してください。\n"
#                     "\n"
#                     "【有効な部品識別子とは】\n"
#                     "- メーカ品番（例: RK73Z1ETTP, CF1/4CS100J, MF1/2CC1003F, NV73DL1JTTE47）\n"
#                     "- Hコード（例: H123456）\n"
#                     "- 品番は英数字と記号（/, -, .）を含み、通常5文字以上\n"
#                     "\n"
#                     "【絶対に抽出しないもの】\n"
#                     "- 内部管理コード（例: EC00384035, 9KJ11105000, TD0-14524001, TE0-06366001）\n"
#                     "- 日本語品名のみ（例: 抵抗器）\n"
#                     "- 金額、納期、税区分、備考、ヘッダ、フッタ、合計行\n"
#                     "\n"
#                     "【数量ルール】\n"
#                     "- 数量は対応する「数量」欄の数値（例: 100, 20000）\n"
#                     "- 不明な場合は「1」\n"
#                     "\n"
#                     "【出力形式】\n"
#                     "- 1行につき「部品識別子,数量」（カンマ区切り）\n"
#                     "- 例:\n"
#                     "  RK73Z1ETTP,100\n"
#                     "  H123456,5\n"
#                     "  CF1/4CS100J,20000\n"
#                     "- 説明文、マークダウン、空白行、引用符は含めない\n"
#                     "- 純粋なテキストのみ出力

#                 """
                
#                 response = client.chat.completions.create(
#                     model="gpt-4o-mini",
#                     messages=[{
#                         "role": "user",
#                         "content": [
#                             {"type": "text", "text": prompt},
#                             {"type": "image_url", "image_url": {"url": image_data_url}}
#                         ]
#                     }],
#                     max_tokens=1500,
#                     temperature=0.1
#                 )
                
#                 page_text = getattr(response.choices[0].message, 'content', "") or ""
#                 if page_text:
#                     all_ocr_text += page_text + "\n"
#                     log.info(f"✅ Page {page_num} OCR completed: {len(page_text)} characters")
#                     log.info(f"📄 OCR text sample: {page_text[:200]}...")
                
#             except Exception as page_error:
#                 log.error(f"❌ Error processing page {page_num}: {page_error}", exc_info=True)
#             finally:

#                 try:
#                     os.remove(img_path)
#                 except Exception:
#                     pass
        
#         log.info(f"📄 OCR completed for {len(images)} pages, total text: {len(all_ocr_text)} characters")
#         return all_ocr_text.strip()
        
#     except Exception as e:
#         log.error(f"❌ OCR error: {e}", exc_info=True)
#         return ""





# def extract_items_from_attachment(file_path):
    
#     if not isinstance(file_path, str) or not os.path.isfile(file_path):
#         log.warning("Invalid file path provided to extract_items_from_attachment")
#         return []

#     items = []
#     filename = os.path.basename(file_path)
#     log.info(f"📎 Processing attachment: {filename}")

#     try:
       
#         if filename.lower().endswith('.pdf'):
#             log.info(f"📄 Processing Japanese PDF attachment: {filename}")
#             with open(file_path, 'rb') as f:
#                 pdf_bytes = f.read()
#             ocr_text = ocr_pdf_with_openai(pdf_bytes)
#             if not ocr_text:
#                 log.warning(" No text extracted from PDF via OCR")
#                 return items
#             log.info(f" OCR extracted {len(ocr_text)} characters")
#             log.info(f" Full OCR text:\n{ocr_text}")
#             for line in ocr_text.splitlines():
#                 line = line.strip()
#                 if not line:
#                     continue
#                 log.debug(f"📄 OCR line: {line}")
#                 patterns = [
#                     r'^["\']?([^,\n]+?)["\']?\s*[,，]\s*(\d+)',
#                     r'^([^\s,]+)\s+(\d+)(?:\s*[個本枚セット])?',
#                     r'^([^\s×*]+)\s*[×*]\s*(\d+)',
#                     r'^([^\s]+)\s+(\d+)\s*[個本枚]',
#                     r'^(H\d{6})\s*[:：]?\s*(\d+)',
#                     r'^([A-Z0-9\u3000-\u9FFF\-_]+)\s+(\d+)\s*[^\d\s]*$',
#                 ]
#                 part_num = None
#                 qty = 1
#                 for pattern in patterns:
#                     m = re.match(pattern, line, re.IGNORECASE)
#                     if m:
#                         part_num = m.group(1).strip()
#                         try:
#                             qty = int(m.group(2))
#                         except (ValueError, TypeError):
#                             qty = 1
#                         log.info(f"📦 Pattern matched: {part_num} x {qty}")
#                         break
#                 if not part_num:
#                     h_match = re.search(r'(H\s*\d{6})', line, re.IGNORECASE)
#                     if h_match:
#                         part_num = h_match.group(1).replace(" ", "").upper()
#                         qty_match = re.search(r'H\s*\d{6}\s*[×*]?\s*(\d+)', line, re.IGNORECASE)
#                         if qty_match:
#                             try:
#                                 qty = int(qty_match.group(1))
#                             except:
#                                 qty = 1
#                     else:
#                         tokens = re.findall(r'[A-Z0-9\u3000-\u9FFF\-_]{2,}', line)
#                         if tokens:
#                             part_num = tokens[0]
#                             qty_match = re.search(r'(\d+)\s*[個本枚]?', line)
#                             if qty_match:
#                                 try:
#                                     qty = int(qty_match.group(1))
#                                 except:
#                                     qty = 1
#                 if part_num:
#                     part_num = part_num.strip().replace(" ", "").upper()
#                     part_num = re.sub(r'[\(（].*[\)）]', '', part_num)
#                     part_num = part_num.strip()
#                     if len(part_num) >= 2:
#                         items.append((part_num, qty))
#                         log.info(f"📦 Final extracted: {part_num} x {qty}")

#         # --- CSV PROCESSING ---

#         elif filename.lower().endswith('.csv'):
#             log.info(f"📊 Processing CSV attachment: {filename}")
#             with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
#                 sample = f.read(2048)
#                 f.seek(0)
#                 try:
#                     delimiter = csv.Sniffer().sniff(sample).delimiter
#                 except Exception:
#                     delimiter = ','
#                 reader = csv.DictReader(f, delimiter=delimiter)
#                 for row in reader:
#                     row_lower = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items() if k is not None}
#                     qty_val = (row_lower.get('qty') or row_lower.get('quantity') or 
#                               row_lower.get('数量') or row_lower.get('個数') or 
#                               row_lower.get('q') or '1')
#                     try:
#                         qty = int(float(qty_val))
#                     except:
#                         qty = 1
#                     part_val = (row_lower.get('hcod') or row_lower.get('品番') or 
#                                row_lower.get('hnm') or row_lower.get('part number') or 
#                                row_lower.get('mpn') or row_lower.get('品名') or None)
#                     if part_val and str(part_val).strip():
#                         clean_part = str(part_val).strip().strip('"').strip("'")
#                         items.append((clean_part, qty))
#                         log.info(f"📦 CSV found: {clean_part} x {qty}")
#                     else:
#                         for k, v in row_lower.items():
#                             if k in ['qty', 'quantity', '数量', '個数', 'q', '']:
#                                 continue
#                             if v and v.strip():
#                                 clean_part = v.strip().strip('"').strip("'")
#                                 items.append((clean_part, qty))
#                                 log.info(f"📦 CSV found (from column {k}): {clean_part} x {qty}")
#                                 break

#         # --- TEXT FILE PROCESSING ---
#         elif filename.lower().endswith(('.txt', '.text')):
#             log.info(f"📝 Processing text attachment: {filename}")
#             with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
#                 txt = f.read()
#             txt_items = extract_items_from_text(txt)
#             items.extend(txt_items)
#             log.info(f"📦 TXT found {len(txt_items)} items")

#         # --- EXCEL FILE PROCESSING ---

#         elif filename.lower().endswith(('.xlsx', '.xls')):
#             log.info(f"📈 Processing Excel attachment: {filename}")
#             try:
#                 import openpyxl
#                 workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
#                 sheet = workbook.active
#                 excel_items = []
#                 for row in sheet.iter_rows(values_only=True):
#                     row_text = ' '.join(str(cell) for cell in row if cell is not None)
#                     if row_text.strip():
#                         row_items = extract_items_from_text(row_text)
#                         excel_items.extend(row_items)
#                 items.extend(excel_items)
#                 log.info(f"📦 Excel found {len(excel_items)} items")
#                 workbook.close()
#             except Exception as e:
#                 log.warning(f"Excel processing error: {e}")
#                 try:
#                     with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
#                         txt = f.read()
#                     fallback_items = extract_items_from_text(txt)
#                     items.extend(fallback_items)
#                     log.info(f"📦 Excel fallback found {len(fallback_items)} items")
#                 except Exception as fallback_error:
#                     log.error(f"Excel fallback also failed: {fallback_error}")

#         # --- OTHER FILE TYPES ---

#         else:
#             log.info(f"🔍 Processing other file type: {filename}")
#             try:
#                 with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
#                     txt = f.read()
#                 other_items = extract_items_from_text(txt)
#                 items.extend(other_items)
#                 log.info(f"📦 Other file type found {len(other_items)} items")
#             except Exception:
#                 log.info(f"❌ Skipping unknown/binary attachment type: {filename}")

#         log.info(f"📦 Total items extracted from attachment: {len(items)}")
#         return items

#     except Exception as e:
#         log.error(f"❌ Attachment processing error ({file_path}): {e}", exc_info=True)
#         return []
    
# def correct_ocr_code(ocr_code: str, known_parts: list):
#     if not ocr_code:
#         return ocr_code
#     code = str(ocr_code).strip().upper()

   
#     if len(code) < 4:
#         return code

#     original = code

#     pattern_corrections = [
#         (" ", ""),        
#         ("O", "0"),       
#         ("I", "1"),      
#         ("L", "1"),       
#         ("B1JTD", "B1JTTD"),
#     ]
#     for wrong, right in pattern_corrections:
#         if wrong in code:
#             code = code.replace(wrong, right)

   
#     try:
#         if known_parts:
#             match = get_close_matches(code, known_parts, n=1, cutoff=0.82)
#             if match:
#                 corrected = match[0].strip().upper()
#                 if corrected != original:
#                     log.info(f"[OCR Correction] '{original}' → '{corrected}' (fuzzy match)")
#                 return corrected
#     except Exception as e:
#         log.debug(f"OCR fuzzy match error: {e}")

#     if code != original:
#         log.info(f"[OCR Correction] '{original}' → '{code}' (pattern fix)")
#     return code





# # for uploading file

# def process_uploaded_file_for_items(filepath: str):
    
#     filename = os.path.basename(filepath)
#     items = []
#     try:
#         if filename.lower().endswith('.pdf'):
#             with open(filepath, 'rb') as f:
#                 pdf_bytes = f.read()
#             ocr_text = ocr_pdf_with_openai(pdf_bytes)
#             log.debug(f"OCR raw output for uploaded file: {repr(ocr_text)}")

#             for line in (ocr_text or "").splitlines():
#                 line = line.strip()
#                 if not line:
#                     continue
#                 m = re.match(r'^["\']?([^\n,]+?)["\']?\s*[,，]\s*(\d+)', line)
#                 if m:
#                     part_num = m.group(1).strip()
#                     qty = int(m.group(2))
#                     items.append({"hcod": part_num, "qty": qty})
#                 else:
#                     token = line.strip().strip('"').strip("'")
#                     if token:
#                         items.append({"hcod": token, "qty": 1})

#         elif filename.lower().endswith('.csv'):
#             with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
#                 sample = f.read(2048)
#                 f.seek(0)
#                 try:
#                     delimiter = csv.Sniffer().sniff(sample).delimiter
#                 except Exception:
#                     delimiter = ','
#                 reader = csv.DictReader(f, delimiter=delimiter)
#                 for row in reader:
#                     row_lower = { (k or "").strip().lower(): (v or "").strip() for k, v in row.items() if k is not None }
                    
                    
#                     qty_val = row_lower.get('qty') or row_lower.get('quantity') or row_lower.get('数量') or row_lower.get('q') or '1'
#                     try:
#                         qty = int(float(qty_val))
#                     except:
#                         qty = 1

                    
#                     part_val = row_lower.get('hcod') or row_lower.get('品番') or row_lower.get('hnm') or \
#                                row_lower.get('part number') or row_lower.get('mpn') or None
#                     if part_val and str(part_val).strip():
#                         items.append({"hcod": str(part_val).strip().strip('"').strip("'"), "qty": qty})
#                     else:
#                         # fallback: first non-qty field
#                         for k, v in row_lower.items():
#                             if k in ['qty', 'quantity', '数量', 'q', '']:
#                                 continue
#                             if v and v.strip():
#                                 items.append({"hcod": v.strip().strip('"').strip("'"), "qty": qty})
#                                 break

#         elif filename.lower().endswith(('.txt', '.text')):
#             with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
#                 txt = f.read()
#             extracted_items = extract_items_from_text(txt)
           
#             for hcod, qty in extracted_items:
#                 items.append({"hcod": hcod, "qty": qty})

#         else:
           
#             with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
#                 txt = f.read()
#             extracted_items = extract_items_from_text(txt)
#             for hcod, qty in extracted_items:
#                 items.append({"hcod": hcod, "qty": qty})

#     except Exception as e:
#         log.warning(f"Attachment error ({filename}): {e}", exc_info=True)

#     return items








# backend/utils/ocr_utils.py
import tempfile
import base64
import os
import re
import logging
from pdf2image import convert_from_bytes
from config import Config
from openai import OpenAI
from difflib import get_close_matches
import csv

# Optional: pdfplumber for structured parsing
try:
    import pdfplumber
    from io import BytesIO
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    logging.getLogger(__name__).warning("pdfplumber not installed. Falling back to AI OCR.")

client = OpenAI(api_key=Config.OPENAI_API_KEY) if getattr(Config, "OPENAI_API_KEY", None) else None

log = logging.getLogger(__name__)


def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r'[０-９]', lambda m: str(ord(m.group()) - 0xFEE0), text)
    return re.sub(r'\s+', ' ', text.strip())


def extract_items_from_text(text: str):
    """Fallback for non-PDF or unstructured files."""
    if not text:
        return []
    text = normalize_text(text)
    items = []

    hcod_pattern = re.compile(
        r'H\s*(\d{6})\s*(?:[:：\-\s]*[x×*]?\s*(?:qty|数量|個|pcs|pc|ea|本|枚|セット|units?)?\s*[:：]?\s*)?(\d+)',
        re.IGNORECASE
    )
    for m in hcod_pattern.finditer(text):
        code = f"H{m.group(1)}"
        qty = int(m.group(2))
        items.append((code, qty))

    all_hcods = re.findall(r'H\s*(\d{6})', text, re.IGNORECASE)
    existing = {code for code, _ in items}
    for digits in all_hcods:
        code = f"H{digits}"
        if code not in existing:
            items.append((code, 1))

    lines = re.split(r'[\n\r]+', text)
    for line in lines:
        line = line.strip()
        if not line:
            continue
        part_qty = re.search(
            r'([\w\s\-–―/／\(\)\[\]\u3000-\u9FFFΩμa-zA-Z0-9\.\+\=\&]{2,80}?)'
            r'\s*[x×＊*]\s*'
            r'(\d{1,5})'
            r'(?:\s*(?:pcs|個|本|枚|セット|pc|ea|units?|点))?\b',
            line,
            re.IGNORECASE
        )
        if part_qty:
            name = part_qty.group(1).strip()
            qty = int(part_qty.group(2))
            name = re.sub(r'[,\.\-_:\s;、。]+$', '', name)
            if len(name) >= 2:
                items.append((name, qty))
            continue
        fallback = re.search(
            r'([\w\s\-–―/／\(\)\[\]\u3000-\u9FFFΩμa-zA-Z0-9\.\+\=\&]{2,80}?)\s+(\d{1,4})\s*$',
            line,
            re.IGNORECASE
        )
        if fallback:
            name = fallback.group(1).strip()
            qty = int(fallback.group(2))
            name = re.sub(r'[,\.\-_:\s;、。]+$', '', name)
            if len(name) >= 2:
                items.append((name, qty))
    return items


def extract_items_from_structured_po(pdf_bytes: bytes):
    """
    Handles both:
    - Tokyo Tron (paired rows, lined table)
    - Futami Denki (simple borderless table)
    """
    if not PDFPLUMBER_AVAILABLE:
        return None

    items = []
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                # Try line-based first (Tokyo Tron)
                table = page.extract_table({
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines"
                })

                # If too few rows, try text-based (Futami Denki)
                if not table or len(table) <= 3:
                    table = page.extract_table({
                        "vertical_strategy": "text",
                        "horizontal_strategy": "text",
                        "min_words_vertical": 2,
                        "min_words_horizontal": 2
                    })

                if not table or len(table) < 2:
                    continue

                # Detect table type by checking for paired structure
                is_paired = False
                for i in range(1, min(5, len(table))):
                    row = table[i]
                    if row and len(row) >= 2:
                        cell = (row[1] or "").strip()
                        if re.match(r'^(9KJ|TD0-|TE0-|EC\d{8})', cell):
                            is_paired = True
                            break

                if is_paired:
                    # Tokyo Tron style: paired rows
                    rows = table[2:] if len(table) > 2 else table
                    i = 0
                    while i < len(rows) - 1:
                        row_a = rows[i]      # internal code row (has quantity)
                        row_b = rows[i + 1]  # real part row

                        if len(row_a) >= 5 and len(row_b) >= 2:
                            qty_field = (row_a[4] or "").strip()
                            qty_str = qty_field.split('\n')[0].replace(',', '').strip()
                            qty = int(qty_str) if qty_str.isdigit() else 1

                            part_name = (row_b[1] or "").strip()
                            if part_name and re.match(r'^[A-Za-z0-9/\.\-_]{5,}$', part_name):
                                if not re.match(r'^(9KJ|TD0-|TE0-|EC\d{8})', part_name):
                                    items.append((part_name, qty))
                        i += 2
                else:
                    # Futami Denki style: simple table
                    for row in table[1:]:
                        if not row or len(row) < 4:
                            continue
                        part_name = None
                        qty = 1

                        # Find part number (alphanum + dash, e.g., 36110-3000FD)
                        for cell in row:
                            cell = (cell or "").strip()
                            if re.match(r'^[A-Za-z0-9\-]{6,}$', cell):
                                part_name = cell
                                break

                        # Find quantity (small integer)
                        for cell in row:
                            cell = (cell or "").strip()
                            if cell.isdigit() and 1 <= int(cell) <= 99999:
                                qty = int(cell)
                                break

                        if part_name:
                            items.append((part_name, qty))

        log.info(f"✅ Extracted {len(items)} items from PDF")
        return items

    except Exception as e:
        log.warning(f"Structured parsing failed: {e}", exc_info=True)
        return None


def ocr_pdf_with_openai(pdf_bytes: bytes) -> str:
    """Fallback OCR using GPT-4o Vision."""
    if not client:
        log.warning("OpenAI client not configured; skipping OCR.")
        return ""

    try:
        log.info("🔄 Converting PDF to image for fallback OCR...")
        images = convert_from_bytes(pdf_bytes, first_page=1, last_page=3)
        if not images:
            return ""

        all_ocr_text = ""
        for page_num, image in enumerate(images, 1):
            img_path = os.path.join(tempfile.gettempdir(), f"temp_ocr_page_{page_num}.jpg")
            try:
                image.save(img_path, format="JPEG", quality=85)
                with open(img_path, "rb") as f:
                    encoded_img = base64.b64encode(f.read()).decode('utf-8')
                image_data_url = f"data:image/jpeg;base64,{encoded_img}"

                prompt = (
                    "以下の注文書画像から、有効な「部品識別子」と「数量」のみを抽出してください。\n"
                    "【有効な部品識別子とは】\n"
                    "- メーカ品番（例: RK73Z1ETTP, CF1/4CS100J, 36110-3000FD）\n"
                    "- Hコード（例: H123456）\n"
                    "- 品番は英数字と記号（/, -, .）を含み、通常5文字以上\n"
                    "【絶対に抽出しないもの】\n"
                    "- 内部管理コード（例: EC00384035, 9KJ11105000, TD0-14524001）\n"
                    "- 日本語品名のみ（例: 抵抗器）\n"
                    "- 金額、納期、税区分、備考、ヘッダ、フッタ、合計行\n"
                    "【数量ルール】\n"
                    "- 数量は対応する「数量」欄の数値\n"
                    "- 不明な場合は「1」\n"
                    "【出力形式】\n"
                    "- 1行につき「部品識別子,数量」（カンマ区切り）\n"
                    "- 例: RK73Z1ETTP,100\n"
                    "- 純粋なテキストのみ出力"
                )

                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": image_data_url}}
                        ]
                    }],
                    max_tokens=1500,
                    temperature=0.1
                )
                page_text = getattr(response.choices[0].message, 'content', "") or ""
                if page_text:
                    all_ocr_text += page_text + "\n"
            except Exception as e:
                log.error(f"Page {page_num} OCR error: {e}", exc_info=True)
            finally:
                try:
                    os.remove(img_path)
                except:
                    pass
        return all_ocr_text.strip()

    except Exception as e:
        log.error(f"OCR fallback failed: {e}", exc_info=True)
        return ""


def extract_items_from_attachment(file_path):
    if not isinstance(file_path, str) or not os.path.isfile(file_path):
        log.warning("Invalid file path")
        return []

    items = []
    filename = os.path.basename(file_path)
    log.info(f"📎 Processing: {filename}")

    try:
        if filename.lower().endswith('.pdf'):
            log.info("📄 Parsing PDF with hybrid method")
            with open(file_path, 'rb') as f:
                pdf_bytes = f.read()

            items = extract_items_from_structured_po(pdf_bytes) if PDFPLUMBER_AVAILABLE else None
            if not items:
                log.info("🔄 Falling back to GPT-4o Vision OCR")
                ocr_text = ocr_pdf_with_openai(pdf_bytes)
                if ocr_text:
                    for line in ocr_text.splitlines():
                        line = line.strip()
                        if not line:
                            continue
                        m = re.match(r'^["\']?([^,\n]+?)["\']?\s*[,，]\s*(\d+)', line)
                        if m:
                            items.append((m.group(1).strip(), int(m.group(2))))
                        else:
                            items.append((line, 1))
            log.info(f"📦 Extracted {len(items)} items")
            return items

        # --- Other file types (unchanged) ---
        elif filename.lower().endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                sample = f.read(2048)
                f.seek(0)
                try:
                    delimiter = csv.Sniffer().sniff(sample).delimiter
                except:
                    delimiter = ','
                reader = csv.DictReader(f, delimiter=delimiter)
                for row in reader:
                    row_lower = { (k or "").strip().lower(): (v or "").strip() for k, v in row.items() if k is not None }
                    qty_val = row_lower.get('qty') or row_lower.get('quantity') or row_lower.get('数量') or row_lower.get('個数') or row_lower.get('q') or '1'
                    qty = int(float(qty_val)) if qty_val.replace('.', '').isdigit() else 1
                    part_val = row_lower.get('hcod') or row_lower.get('品番') or row_lower.get('hnm') or row_lower.get('part number') or row_lower.get('mpn') or None
                    if part_val and str(part_val).strip():
                        items.append((str(part_val).strip().strip('"').strip("'"), qty))
                    else:
                        for k, v in row_lower.items():
                            if k in ['qty', 'quantity', '数量', '個数', 'q', '']:
                                continue
                            if v and v.strip():
                                items.append((v.strip().strip('"').strip("'"), qty))
                                break

        elif filename.lower().endswith(('.txt', '.text')):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                txt = f.read()
            items = extract_items_from_text(txt)

        elif filename.lower().endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' '.join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        items.extend(extract_items_from_text(row_text))
                workbook.close()
            except Exception as e:
                log.warning(f"Excel error: {e}")
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        txt = f.read()
                    items = extract_items_from_text(txt)
                except:
                    pass

        else:
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    txt = f.read()
                items = extract_items_from_text(txt)
            except:
                log.info(f"Skipping binary file: {filename}")

        log.info(f"📦 Total items: {len(items)}")
        return items

    except Exception as e:
        log.error(f"Attachment error: {e}", exc_info=True)
        return []


def correct_ocr_code(ocr_code: str, known_parts: list):
    if not ocr_code:
        return ocr_code
    code = str(ocr_code).strip().upper()
    if len(code) < 4:
        return code

    original = code
    pattern_corrections = [(" ", ""), ("O", "0"), ("I", "1"), ("L", "1"), ("B1JTD", "B1JTTD")]
    for wrong, right in pattern_corrections:
        code = code.replace(wrong, right)

    try:
        if known_parts:
            match = get_close_matches(code, known_parts, n=1, cutoff=0.82)
            if match:
                corrected = match[0].strip().upper()
                if corrected != original:
                    log.info(f"[OCR Correction] '{original}' → '{corrected}'")
                return corrected
    except Exception as e:
        log.debug(f"Fuzzy match error: {e}")

    if code != original:
        log.info(f"[OCR Correction] '{original}' → '{code}'")
    return code


def process_uploaded_file_for_items(filepath: str):
    filename = os.path.basename(filepath)
    items = []
    try:
        if filename.lower().endswith('.pdf'):
            with open(filepath, 'rb') as f:
                pdf_bytes = f.read()
            structured_items = extract_items_from_structured_po(pdf_bytes) if PDFPLUMBER_AVAILABLE else None
            if structured_items:
                for part, qty in structured_items:
                    items.append({"hcod": part, "qty": qty})
            else:
                ocr_text = ocr_pdf_with_openai(pdf_bytes)
                for line in (ocr_text or "").splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    m = re.match(r'^["\']?([^\n,]+?)["\']?\s*[,，]\s*(\d+)', line)
                    if m:
                        items.append({"hcod": m.group(1).strip(), "qty": int(m.group(2))})
                    else:
                        items.append({"hcod": line, "qty": 1})
        else:
            raw_items = extract_items_from_attachment(filepath)
            for part, qty in raw_items:
                items.append({"hcod": part, "qty": qty})
    except Exception as e:
        log.warning(f"File processing error: {e}", exc_info=True)
    return items